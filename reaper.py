
import os
import zipfile
import tempfile
import openpyxl
import shutil
from itertools import chain
import datetime
import pprint


#пашет в Linux
def fix_xlsx_2(in_file):
    '''
    приводит xlsx файл из 1С к стандартной структуре
    '/xl/sharedStrings.xml' -> '/xl/SharedStrings.xml'
    :param in_file:
    :return:
    '''
    tmp_dir = os.path.curdir+'/tmp'
   
    if os.path.exists(tmp_dir):
        shutil.rmtree(tmp_dir)
    os.mkdir(tmp_dir)    
    with zipfile.ZipFile(in_file, 'r') as z_file:
        z_file.extractall(path=tmp_dir)
    os.rename(tmp_dir+'/xl/sharedStrings.xml',tmp_dir+'/xl/SharedStrings.xml')
    os.remove(in_file)
    shutil.make_archive(
        in_file,
        'zip',
        root_dir=tmp_dir
    )
    os.rename('{}.zip'.format(in_file), in_file)

# не пашет в Linux
def fix_xlsx(in_file):
    '''
    приводит xlsx файл из 1С к стандартной структуре
    '/xl/sharedStrings.xml' -> '/xl/SharedStrings.xml'
    :param in_file:
    :return:
    '''
    print('test')
    print(os.path.curdir)
    tmpfd, tmp = tempfile.mkstemp(dir=os.path.dirname(in_file))
    #print(tmp)
    #print(tmpfd)
    os.close(tmpfd)
    _filename = '[Content_Types].xml'
    data = ''
    with zipfile.ZipFile(in_file, 'r') as zin:
        with zipfile.ZipFile(tmp, 'w') as zout:
            for item in zin.infolist():
                print(item)
                if item.filename != _filename:
                    zout.writestr(item, zin.read(item.filename))
                else:
                    data = zin.read(_filename).decode()
    os.remove(in_file)
    os.rename(tmp, in_file)
    data = data.replace('/xl/sharedStrings.xml', '/xl/SharedStrings.xml')
    
    with zipfile.ZipFile(in_file, mode='a', compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(_filename, data)


class XLSX_reaper:

#   Структура словаря
# 1    _id : _id	
# 2   nativeName : nativeName	
# 3   fullAddress : fullAddress	
# 4   coordinates : coordinates   
# 5   photo_url : photo_url
# 6   create_date : create_date
# 7   object_type : object_type
# 8   category_type : category_type
# 9   reg_number : reg_number 
# 10  con_number : con_number
# 11  documents : documents
# 12  addr_NPA : Адрес в соответствии с НПА
# 13   municipality : Муниципальное образование	
# 14   locality : Населенный пункт	
# 15   OKN_in_ensemble : ОКН входит в ансамбль (да/нет)	
# 16   information_sign : Наличие утановленной информационной надписи установленного образца да/нет/на подготовке	
# 17   information_sign_photo : Фотография информационной надписи	
# 18   information_sign_conformity  Наличие информационной надписи, но не соответствующей требованиям  (да/нет)	
# 19   photo : Фото	
# 20   url : URL-адрес	
# 21   descripton : Описание	
# 22   affiliation_U : Принадлежность к ЮНЭСКО  (да/нет)
# 23   esp_valuable_object :  Особо ценный объект  (да/нет)
# 24   requisites_and_title : Реквизиты и наименование акта органа государственной власти о постановке на государственную охрану объекта культурного наследия	
# 25   owner : Собственник	
# 26   management Под чьим управлением	
# 27   owner_contacts : Контактные данные собственника ОКН	
# 28   has_security_obligation : Наличие охранного обязательства ОКН  да/нет/на подготовке	
# 29   has_passport_OKN : Наличие паспорта ОКН да/нет/на подготовке	
# 30   actual_address : Актуальный адрес	
# 31   gen_species_appearance : Общая видовая принадлежность	
# 32   has_docs_boundaries : Наличие документов о границах территории ОКН	
# 33   req_of_approval : Реквизиты об утверждении границ территории	
# 34   has_docs_of_aprroval : Наличие документов об утвержденых зонах охраны	
# 35   document_on_approved_security Документ об утвержденых зонах охраны	
# 36   has_rights : Наличие зарегистрированных прав/обременений  (да/нет)	
# 37   date : Дата


    def __init__(self, filename):
        self.filename = filename
        self.__first_row = 2
        self.result_list = []


    def __make_dict(self, row):
        result = dict() 
        result['_id'] = row[0] #1
        result['nativeName'] = row[1] #2
        result['fullAddress'] = row[2] #3
        result['coordinates'] = row[3] #4
        result['photo_url'] = row[4] #5
        result['create_date'] = row[5] #6
        result['object_type'] = row[6] #7
        result['category_type'] = row[7] #8
        result['reg_number'] = row[8] #9
        result['con_number'] = row[9] #10
        result['documents'] = row[10] #11
        result['addr_NPA'] = row[11] #12
        result['municipality'] = row[12] #13
        result['locality'] = row[13] #14
        result['OKN_in_ensemble'] = row[14] #15
        result['information_sign'] = row[15] #16
        result['information_sign_photo'] = row[16] #17
        result['information_sign_conformity'] = row[17] #18
        result['photo'] = row[18] #19
        result['url'] = row[19] #20
        result['description'] = row[20] #21
        result['affiliation_U'] = row[21] #22
        result['esp_valuable_object'] = row[22] #23
        result['requisites_and_title'] = row[23] #24
        result['owner'] = row[24] #25
        result['management'] = row[25] #26
        result['owner_contacts'] = row[26] #27
        result['has_security_obligation'] = row[27] #28
        result['has_passport_OKN'] = row[28] #29
        result['actual_address'] = row[29] #30
        result['gen_species_appearance'] = row[30] #31
        result['has_docs_boundaries'] = row[31] #32
        result['req_of_approval'] = row[32] #33
        result['has_docs_of_aprroval'] = row[33] #34
        result['document_on_approved_security'] = row[34] #35
        result['has_rights'] = row[35] #36
        result['date'] = row[36] #37
        return result 
        

        
    def _get_items(self):
        # нормализация
        self.result_list.clear()
        try:
            wb = openpyxl.load_workbook(filename=self.filename)
            print('Нормализация XLSX файл не требуется')
        except KeyError:
            print('Нормализация XLSX файла')
            fix_xlsx_2(self.filename)
            print('Готово')
            wb = openpyxl.load_workbook(filename=self.filename)
        except Exception as ex:
            print(ex)
        ########################################################
        sheet = wb.active
        row = self.__first_row
        while True:
            mark_cell = sheet.cell(row=row, column=1).value
            if mark_cell is None:
                break
            excel_row = [str(sheet.cell(row=row, column=col).value) for col in range(1,38)] 
            self.result_list.append(self.__make_dict(excel_row))
            row += 1
        print('Список сформирован')
        return self.result_list



    # множество типов объектов
    @property
    def object_types(self):
        temp_list = [x['object_type'] for x in self._get_items()]
        return sorted(list(set(temp_list)))

    # множество типов категорий
    @property
    def category_types(self):
        temp_list = [x['category_type'] for x in self._get_items()]
        return set(temp_list)

    #список словарей из XLSX
    @property
    def xlsx_items(self):
        return self._get_items()

    # множество муниципальных образований
    @property
    def municipalitys(self):
        temp_list = [x['municipality'] for x in self._get_items()]
        return sorted(list(set(temp_list)))
     
    # множество населенных пунктов
    @property
    def localitys(self):
        temp_list = [x['locality'] for x in self._get_items()]
        return set(temp_list)

    # множество видовых принадлежностей
    @property
    def species(self):
        temp_list = [x['gen_species_appearance'] for x in self._get_items()]
        return set(temp_list)


if __name__ == '__main__':
    xr = XLSX_reaper('file.xlsx')
  
    for d in xr.municipalitys:
        print(d)
  #print(len(xr.xlsx_items))

