import json
from django.core.management.base import BaseCommand, CommandError
from django.apps import apps
import openpyxl
from collections import defaultdict



from dashboard.models import Municipality
from dashboard.models import Locality


class Command(BaseCommand):
    


    def handle(self, *args, **kwargs):
        self.__create_list('geo2.xlsx')
    
    
    def __create_list(self, filename):
        wb = openpyxl.load_workbook(filename=filename)
        sheet = wb.active
        row = 1
        m_list = []
        geo_dict = dict()
        while True:
            mark_cell = sheet.cell(row=row, column=1).value
            if mark_cell is None:
                break
            municipality_str = sheet.cell(row=row, column=2).value
            m_list.append(municipality_str)
            row += 1
        
        #print(m_list)
        m_list = list(set(m_list))
        for m in m_list:
            geo_dict[m] = []
            
                
        row = 1
        while True:
            mark_cell = sheet.cell(row=row, column=1).value
            if mark_cell is None:
                break
            municipality_str = sheet.cell(row=row, column=2).value
            locality_str = sheet.cell(row=row, column=1).value
            if municipality_str is not None:
                geo_dict[municipality_str].append(locality_str)
            row += 1

        print(geo_dict.keys())
        
        #заполнить отсутствующие муниципальные образования
        for m in geo_dict.keys():
            print(m)
            #print(Municipality.objects.filter(name=m))
            if len(Municipality.objects.filter(name=m)) == 0:
                mun = Municipality(name=m)
                mun.save()
            #print("{} {}".format(m, Municipality.objects.filter(name=m)))

        for m in geo_dict.keys():
            #print(m)
            #print(Municipality.objects.filter(name=m))
            print("{} {}".format(m, Municipality.objects.filter(name=m)))

        
        
        total_before = len(Locality.objects.all())
        for municipality in Municipality.objects.all():
            if municipality.name in geo_dict.keys():
                for locality_name in geo_dict[municipality.name]:
                    if len(Locality.objects.filter(name=locality_name)) == 0:
                        print("New locality : {} in {}".format(locality_name, municipality.name))
                        new_locality = Locality(name=locality_name, municipality=municipality)
                        new_locality.save()
        total_after = len(Locality.objects.all())
        print("Total before = {}".format(total_before))
        print("Total after = {}".format(total_after))
        
        
    
